# pyt-able.py
# module for EXCEL exports

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def create_demo(filename="example.xlsx", title="Example"):
    workbook = Workbook()

    try:
        worksheet = workbook.active
        worksheet.title = title

        # merge cells for header
        worksheet.merge_cells('A1:F1')
        worksheet["A1"] = title

        set_columns_width(worksheet, 50, for_all=True)
        set_columns_width(worksheet, (50, 20), for_all=True)
        set_rows_height(worksheet, (20, 10, 20, 15))

    finally:
        workbook.save(filename=filename)


def set_columns_width(worksheet, widths, for_all=False, col_min=1, col_max=16384):
    if col_min > col_max or col_max < 1 or col_min < 1 or col_max > 16384 or col_min > 16384:
        raise AssertionError("Wrong col_min/col_max parameters")

    if type(widths) not in (list, tuple, int):
        if for_all:
            raise AssertionError("widths argument is wrong with for_all (must be list or tuple or int)")
        else:
            raise AssertionError("widths argument is wrong without for_all (must be list or tuple)")

    if for_all:
        if type(widths) in (list, tuple):
            for no in range(col_min, col_max+1):
                worksheet.column_dimensions[get_column_letter(no)].width = widths[no % len(widths)-1]

        elif type(widths) == int:
            for no in range(col_min, col_max + 1):
                worksheet.column_dimensions[get_column_letter(no)].width = widths

    else:
        if type(widths) in (list, tuple):
            for no, width in enumerate(widths, 1):
                worksheet.column_dimensions[get_column_letter(no)].width = width


def set_rows_height(worksheet, heights, for_all=False):
    for no, height in enumerate(heights, 1):
        worksheet.row_dimensions[no].height = height


create_demo()
